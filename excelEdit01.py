"""
기존파일 편집
"""
import openpyxl as op
from openpyxl.styles import Color, Font, PatternFill, Border, Side

wb = op.load_workbook("./temp2/WantedInfo.120767~172894.xlsx")
ws = wb.active
# print(ws.max_row)

"""
# 새로운 폰트 설정
new_font = Font(name="맑은 고딕", bold=True, italic=False, color=Color(rgb="0000FF"))
box = Border(left=Side(border_style='thin', color='000000'),g:q:wq
             right=Side(border_style='thin', color='000000'),
             top=Side(border_style='thin', color='000000'),
             bottom=Side(border_style='thin', color='000000'))

for i in range(2, ws.max_row+1):
    cell = ws["A" + str(i)]
    cell.hyperlink = cell.value
    cell.style = "Hyperlink"
    cell.border = box
    cell.font = new_font
"""

for cell in ws["A"]:
    cell.hyperlink = cell.value

wb.save("./temp2/WantedInfo.120767~172894.xlsx")
